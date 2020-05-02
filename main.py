import csv
import os
import openpyxl as op
from openpyxl.styles import Alignment, Font

f1 = open("F1_exam_timetable_with_roll_numbers.csv", "r")
f2 = open("F2_rooms_capacity_info.csv", "r")
r1 = csv.DictReader(f1)
r2 = csv.DictReader(f2)
COLORS = [
    '00000000', '00FF0000', '0000FF00', '000000FF',             #0-4
    '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF', #5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF', #10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000', #15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF', #20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080', #25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00', #30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF', #35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF', #40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC', #45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699', #50-54
    '00969696', '00003366', '00339966', '00003300', '00333300', #55-59
    '00993300', '00993366', '00333399', '00333333', 'System Foreground', 'System Background' #60-64
]
main_list = [dict(row) for row in r1]
room_limit = [dict(row) for row in r2]
n = len(main_list)
check_list = [0] * n    #If this is zero, Then That row is yet to be processed. If It is not, Then Processing is Done.
file_count = 0
for row in main_list:
    ind = main_list.index(row)  #Index of row
    if check_list[ind] != 0:
        continue
    else:
        temp_list = []          #Contains the required rows (Directly)
        for row1 in main_list:
            if row1["date"] == row["date"] and row1["day"] == row["day"] and row1["shift"] == row["shift"] and row1["roomno"] == row["roomno"]:
                temp_list.append(row1)
                check_list[main_list.index(row1)] = 1
        temp_list.sort(key=lambda data: int(data["allocationdone"]), reverse = True)
        #students = []
        """for row1 in temp_list:
            li = []
            li = row1["rollnolist"].split(",")
            courses.add(li[0][:6])
            if li[-1] == "":
                li.pop()            
            students.append(li)
        courses = list(courses)"""
        file_count += 1
        temp_list.sort(key=lambda x: int(x["allocationdone"]), reverse=True)
        for room in room_limit:
            if room["Room No."] == row["roomno"]:
                room_strength = int(room["Exam Capacity"])
        students_roll_list = [] # This list contains the Roll Numbers in Descending Order of their Strength of Batch.
        # We can directly write this list into required file.
        Total_Students = 0
        for row1 in temp_list:
            li = []
            li = row1["rollnolist"].split(",")
            if li[-1] == "":
                li.pop()
            students_roll_list.extend(li)
            Total_Students+=len(li)
        #students_roll_list  = list(set(students_roll_list))
        size = len(students_roll_list)
        final = [" "] * size
        if size%2 == 1:   # If Total Allocation is odd
            k = (size + 1)//2
            for i in range(0, size, 2):
                final[i] = students_roll_list[i//2]
                if i <= size-3:
                    final[i+1] = students_roll_list[(i//2) + k]
        else:                           # If Total Allocation is even
            k = len(students_roll_list)//2
            for i in range(0, size, 2):
                final[i] = students_roll_list[i//2]
                final[i+1] = students_roll_list[(i//2) + k]
        courses = set()
        for i in range(len(final)):
            courses.add(final[i][:6])
        courses = list(courses)
        if size < 4:  
            col_len = size               # Allocation < 4
            c1 = final
            c2 = []
            c3 = []
            c4 = []
        else:                               # Allocation >= 4
            rem = size%4
            if rem != 0:
                size += (4-rem)
            col_len = size//4
            c1 = [] * col_len
            c2 = [] * col_len
            c3 = [] * col_len
            c4 = [] * (col_len + rem - 4)
            c1 = final[0:col_len]
            c2 = final[col_len : 2*col_len]
            c3 = final[2*col_len:3*col_len]
            c4 = final[3*col_len:]
        # CREATING AND ADDING DETAILS TO FILE

        folder_path = "Output/" + "_".join(row["date"].split("/")) + "/" + row["shift"]
        if os.path.exists(folder_path) == False:
            os.makedirs(folder_path)
        file_path  = folder_path + "/R" + row["roomno"] + ".xlsx"
        wb = op.Workbook()
        sheet = wb.active
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        # Merge Cells
        sheet.merge_cells("A1:D1")
        sheet.merge_cells("A2:D2")
        sheet.merge_cells("A3:D3")
        heading_font = Font(name='Calibri', size=12, bold=True, color='FF000000', underline='single')
        cell_Alignment = Alignment(horizontal='center', vertical='bottom')
        sub_headings_font = Font(name='Calibri', size=12, bold=True, color='FF000000')
        normal_cell_font = Font(name='Calibri', size=12, color='FF000000')
        sheet["A1"].font = heading_font
        sheet["A1"].alignment = cell_Alignment
        sheet["A1"] = "Seating Chart"
        sheet["A2"] = "                                                                                                 Date: " + ".".join(row["date"].split("/"))
        sheet["A2"].font = sub_headings_font
        sheet["A2"].alignment = cell_Alignment
        sheet["A4"] = "C1"
        sheet["A4"].font = sub_headings_font
        sheet["A4"].alignment = cell_Alignment
        sheet["B4"] = "C2"
        sheet["B4"].font = sub_headings_font
        sheet["B4"].alignment = cell_Alignment
        sheet["C4"] = "C3"
        sheet["C4"].font = sub_headings_font
        sheet["C4"].alignment = cell_Alignment
        sheet["D4"] = "C4"
        sheet["D4"].font = sub_headings_font
        sheet["D4"].alignment = cell_Alignment
        if row["shift"] == "Morning":
            time_lower_limit = "10"
            time_upper_limit = "12"
            mer = "A"   #Meridian
        else:
            time_lower_limit = "03"
            time_upper_limit = "05"
            mer = "P"   #Meridian
        sheet["A3"] = "Room No." + row["roomno"] + "                                                          Time:   "+ time_lower_limit + ":00" + mer + "M - " + time_upper_limit + ":00" + mer + "M"
        sheet["A3"].font = sub_headings_font
        sheet["A3"].alignment = cell_Alignment
        # Filling Sheet START
        #Column 1
        for i in range(len(c1)):
            sheet.cell(row=i+5, column=1).value = c1[i]
            color_index = courses.index(c1[i][:6])
            sheet.cell(row=i+5, column=1).font = Font(name='Calibri', size=12, color=COLORS[color_index])
            sheet.cell(row=i+5, column=1).alignment = cell_Alignment            

        #Column 2
        for i in range(len(c2)):
            sheet.cell(row=i+5, column=2).value = c2[i]
            color_index = courses.index(c2[i][:6])
            sheet.cell(row=i+5, column=2).font = Font(name='Calibri', size=12, color=COLORS[color_index])
            sheet.cell(row=i+5, column=2).alignment = cell_Alignment
        #Column 3
        for i in range(len(c3)):
            sheet.cell(row=i+5, column=3).value = c3[i]
            color_index = courses.index(c3[i][:6])
            sheet.cell(row=i+5, column=3).font = Font(name='Calibri', size=12, color=COLORS[color_index])
            sheet.cell(row=i+5, column=3).alignment = cell_Alignment
        #Column 4
        for i in range(len(c4)):
            sheet.cell(row=i+5, column=4).value = c4[i]
            color_index = courses.index(c4[i][:6])
            sheet.cell(row=i+5, column=4).font = Font(name='Calibri', size=12, color=COLORS[color_index])
            sheet.cell(row=i+5, column=4).alignment = cell_Alignment
        # Filling Sheet END
        wb.save(file_path)
print("Total Files Created : ", file_count)
print("🙂️ 🙂️ 🙂️")
